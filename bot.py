import asyncio
import os
from datetime import datetime

from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery

from openpyxl import Workbook, load_workbook


API_TOKEN = os.getenv("API_TOKEN")
ADMIN_CHAT_ID = int(os.getenv("ADMIN_CHAT_ID"))
EXCEL_FILE = "feedbacks.xlsx"
bot = Bot(token=API_TOKEN)
dp = Dispatcher()

# состояния пользователей
active_feedback = set()
user_category = {}

# категории для кнопок
CATEGORIES = ["Преподаватели", "Предметы", "Расписание", "Инфраструктура", "Другое"]

# создаём Excel, если его нет
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отзывы"
    ws.append(["Время", "Категория", "Отзыв"])
    wb.save(EXCEL_FILE)

# клавиатура с категориями
def category_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=cat, callback_data=f"cat_{i}")]
            for i, cat in enumerate(CATEGORIES)
        ]
    )

# старт
@dp.message(Command("start"))
async def start_handler(message: Message):
    await message.answer("Бот запущен.")

# новая форма отзыва
@dp.message(Command("feedback"))
async def new_feedback_handler(message: Message):
    user_id = message.from_user.id
    active_feedback.add(user_id)
    await message.answer("Выбери категорию:", reply_markup=category_keyboard())

# выбор категории
@dp.callback_query(lambda c: c.data.startswith("cat_"))
async def category_chosen(callback: CallbackQuery):
    user_id = callback.from_user.id
    if user_id not in active_feedback:
        await callback.answer()
        return

    index = int(callback.data.split("_")[1])
    user_category[user_id] = CATEGORIES[index]

    await callback.message.answer(f"Категория выбрана: {CATEGORIES[index]}\nТеперь напиши отзыв.")
    await callback.answer()

# отмена
@dp.message(Command("cancel"))
async def cancel_handler(message: Message):
    user_id = message.from_user.id
    active_feedback.discard(user_id)
    user_category.pop(user_id, None)
    await message.answer("Отменено.")

# обработка текста отзыва
@dp.message()
async def feedback_handler(message: Message):
    user_id = message.from_user.id
    text = message.text.strip()

    if user_id not in active_feedback:
        return

    if user_id not in user_category:
        await message.answer("Сначала выбери категорию.")
        return

    if text.lower() in ["cancel", "отмена", "стоп"]:
        active_feedback.discard(user_id)
        user_category.pop(user_id, None)
        await message.answer("Отменено.")
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    category = user_category[user_id]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([now, category, text])
    wb.save(EXCEL_FILE)

    await bot.send_message(ADMIN_CHAT_ID, f" {now}\n {category}\n {text}")

    active_feedback.discard(user_id)
    user_category.pop(user_id, None)
    await message.answer("Отзыв отправлен.")

# запуск бота
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
